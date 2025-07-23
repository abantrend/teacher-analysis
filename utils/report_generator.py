import pandas as pd
import numpy as np
from io import BytesIO
import streamlit as st
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from typing import Dict
import tempfile
import os
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

class ReportGenerator:
    """مولد التقارير الشاملة"""
    
    def __init__(self):
        self.passing_grade = 50
        # استخدام خطوط بسيطة تدعم العربية
        self.arabic_font = "Helvetica"
        self.arabic_font_bold = "Helvetica-Bold"
    
    def _format_arabic_text(self, text: str) -> str:
        """تنسيق النص العربي للعرض الصحيح في PDF"""
        # إرجاع النص كما هو - Streamlit Cloud يدعم UTF-8 افتراضياً
        return text
        
    def generate_comprehensive_report(self, df: pd.DataFrame, stats: Dict, grade_ranges: pd.DataFrame) -> str:
        """
        إنتاج تقرير شامل بصيغة Excel
        
        Args:
            df: DataFrame يحتوي على البيانات
            stats: قاموس الإحصائيات
            grade_ranges: DataFrame نطاقات الدرجات
            
        Returns:
            مسار الملف المؤقت للتقرير
        """
        # إنشاء ملف مؤقت
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        temp_filename = temp_file.name
        temp_file.close()
        
        # إنشاء workbook
        wb = openpyxl.Workbook()
        
        # إزالة الورقة الافتراضية
        if wb.active:
            wb.remove(wb.active)
        
        # إنشاء الأوراق
        self._create_summary_sheet(wb, stats)
        self._create_detailed_data_sheet(wb, df)
        self._create_grade_ranges_sheet(wb, grade_ranges)
        self._create_top_students_sheet(wb, df)
        self._create_failing_students_sheet(wb, df)
        self._create_statistics_sheet(wb, df, stats)
        
        # حفظ الملف
        wb.save(temp_filename)
        
        return temp_filename
    
    def generate_pdf_report(self, df: pd.DataFrame, stats: Dict, grade_ranges: pd.DataFrame) -> str:
        """
        إنتاج تقرير شامل بصيغة PDF
        
        Args:
            df: DataFrame يحتوي على البيانات
            stats: قاموس الإحصائيات
            grade_ranges: DataFrame نطاقات الدرجات
            
        Returns:
            مسار الملف المؤقت للتقرير
        """
        # إنشاء ملف مؤقت
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        temp_filename = temp_file.name
        temp_file.close()
        
        # إنشاء المستند
        doc = SimpleDocTemplate(temp_filename, pagesize=A4)
        story = []
        
        # تحديد الأنماط العربية
        styles = getSampleStyleSheet()
        
        # أنماط عربية مخصصة
        arabic_title_style = ParagraphStyle(
            'ArabicTitleStyle',
            fontName=self.arabic_font_bold,
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            rightIndent=0,
            leftIndent=0
        )
        
        arabic_heading_style = ParagraphStyle(
            'ArabicHeadingStyle',
            fontName=self.arabic_font_bold,
            fontSize=14,
            spaceAfter=12,
            alignment=TA_RIGHT,
            rightIndent=0,
            leftIndent=0
        )
        
        arabic_normal_style = ParagraphStyle(
            'ArabicNormalStyle',
            fontName=self.arabic_font,
            fontSize=10,
            alignment=TA_RIGHT,
            rightIndent=0,
            leftIndent=0
        )
        
        # العنوان الرئيسي
        title = Paragraph("Student Grade Analysis Report / تقرير تحليل درجات الطلاب", arabic_title_style)
        story.append(title)
        
        # معلومات التقرير
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        date_info = Paragraph(f"Report Date / تاريخ التقرير: {current_time}", arabic_normal_style)
        story.append(date_info)
        story.append(Spacer(1, 20))
        
        # الإحصائيات الأساسية
        story.append(Paragraph("Basic Statistics / الإحصائيات الأساسية", arabic_heading_style))
        
        stats_data = [
            [self._format_arabic_text('المقياس'), self._format_arabic_text('القيمة')],
            [self._format_arabic_text('إجمالي الطلاب'), str(stats['count'])],
            [self._format_arabic_text('متوسط الدرجات'), f"{stats['mean']:.2f}"],
            [self._format_arabic_text('الوسيط'), f"{stats['median']:.2f}"],
            [self._format_arabic_text('الانحراف المعياري'), f"{stats['std']:.2f}"],
            [self._format_arabic_text('أعلى درجة'), str(stats['max'])],
            [self._format_arabic_text('أقل درجة'), str(stats['min'])],
            [self._format_arabic_text('معدل النجاح'), f"{stats['pass_rate']:.1f}%"],
            [self._format_arabic_text('معدل الرسوب'), f"{stats['fail_rate']:.1f}%"],
            [self._format_arabic_text('عدد الناجحين'), str(stats['passing_count'])],
            [self._format_arabic_text('عدد الراسبين'), str(stats['failing_count'])]
        ]
        
        stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), self.arabic_font_bold),
            ('FONTNAME', (0, 1), (-1, -1), self.arabic_font),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(stats_table)
        story.append(Spacer(1, 20))
        
        # نطاقات الدرجات
        story.append(Paragraph(self._format_arabic_text("توزيع نطاقات الدرجات"), arabic_heading_style))
        
        grade_data = [[self._format_arabic_text('النطاق'), self._format_arabic_text('عدد الطلاب'), self._format_arabic_text('النسبة المئوية')]]
        for _, row in grade_ranges.iterrows():
            grade_data.append([
                self._format_arabic_text(str(row['النطاق'])), 
                str(row['عدد الطلاب']), 
                f"{row['النسبة %']:.1f}%"
            ])
        
        grade_table = Table(grade_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
        grade_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), self.arabic_font_bold),
            ('FONTNAME', (0, 1), (-1, -1), self.arabic_font),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(grade_table)
        story.append(PageBreak())
        
        # أفضل الطلاب
        story.append(Paragraph(self._format_arabic_text("أفضل 10 طلاب"), arabic_heading_style))
        top_students = df.nlargest(10, 'النسبة المئوية' if 'النسبة المئوية' in df.columns else 'الدرجة')
        
        top_data = [[self._format_arabic_text('المرتبة'), self._format_arabic_text('اسم الطالب'), self._format_arabic_text('الدرجة')]]
        if 'النسبة المئوية' in df.columns:
            top_data[0].append(self._format_arabic_text('النسبة المئوية'))
            
        for i, (_, student) in enumerate(top_students.iterrows(), 1):
            row = [str(i), self._format_arabic_text(str(student['اسم الطالب'])), str(student['الدرجة'])]
            if 'النسبة المئوية' in df.columns:
                row.append(f"{student['النسبة المئوية']:.1f}%")
            top_data.append(row)
        
        col_widths = [0.8*inch, 2.5*inch, 1*inch]
        if 'النسبة المئوية' in df.columns:
            col_widths.append(1.2*inch)
            
        top_table = Table(top_data, colWidths=col_widths)
        top_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.green),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), self.arabic_font_bold),
            ('FONTNAME', (0, 1), (-1, -1), self.arabic_font),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(top_table)
        story.append(Spacer(1, 20))
        
        # الطلاب المتعثرين
        if 'النسبة المئوية' in df.columns:
            failing_students = df[df['النسبة المئوية'] < 50]
        else:
            failing_students = df[df['الدرجة'] < self.passing_grade]
            
        if not failing_students.empty:
            story.append(Paragraph(self._format_arabic_text("الطلاب المتعثرين"), arabic_heading_style))
            
            fail_data = [[
                self._format_arabic_text('اسم الطالب'), 
                self._format_arabic_text('الدرجة'), 
                self._format_arabic_text('الفجوة'), 
                self._format_arabic_text('الملاحظات')
            ]]
            for _, student in failing_students.iterrows():
                if 'النسبة المئوية' in df.columns:
                    grade_for_note = student['النسبة المئوية']
                    if 'الدرجة الكلية' in df.columns:
                        gap = (student['الدرجة الكلية'] * 0.5) - student['الدرجة']
                    else:
                        gap = 25 - student['الدرجة']
                else:
                    grade_for_note = student['الدرجة']
                    gap = self.passing_grade - student['الدرجة']
                
                if grade_for_note < 30:
                    note = self._format_arabic_text("يحتاج دعم عاجل")
                elif grade_for_note < 40:
                    note = self._format_arabic_text("يحتاج تدخل سريع")
                else:
                    note = self._format_arabic_text("قريب من النجاح")
                
                fail_data.append([
                    self._format_arabic_text(str(student['اسم الطالب'])),
                    str(student['الدرجة']),
                    f"{gap:.1f}",
                    note
                ])
            
            fail_table = Table(fail_data, colWidths=[2*inch, 1*inch, 1*inch, 1.5*inch])
            fail_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.red),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), self.arabic_font_bold),
                ('FONTNAME', (0, 1), (-1, -1), self.arabic_font),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.mistyrose),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(fail_table)
        
        # بناء المستند
        doc.build(story)
        
        return temp_filename
    
    def _create_summary_sheet(self, wb: openpyxl.Workbook, stats: Dict):
        """إنشاء ورقة الملخص العام"""
        ws = wb.create_sheet("الملخص العام", 0)
        
        # تنسيق العنوان الرئيسي
        ws.merge_cells('A1:D1')
        ws['A1'] = "تقرير تحليل درجات الطلاب"
        ws['A1'].font = Font(size=18, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws['A1'].font = Font(size=18, bold=True, color='FFFFFF')
        
        # معلومات التقرير
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A3'] = f"تاريخ إنتاج التقرير: {current_time}"
        ws['A3'].font = Font(size=12, italic=True)
        
        # الإحصائيات الأساسية
        ws['A5'] = "الإحصائيات الأساسية"
        ws['A5'].font = Font(size=14, bold=True)
        ws['A5'].fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        
        stats_data = [
            ['عدد الطلاب', stats['count']],
            ['المتوسط الحسابي', f"{stats['mean']:.2f}"],
            ['الوسيط', f"{stats['median']:.2f}"],
            ['الانحراف المعياري', f"{stats['std']:.2f}"],
            ['أعلى درجة', stats['max']],
            ['أقل درجة', stats['min']],
            ['نسبة النجاح', f"{stats['pass_rate']:.1f}%"],
            ['نسبة الرسوب', f"{stats['fail_rate']:.1f}%"],
            ['عدد الناجحين', stats['passing_count']],
            ['عدد الراسبين', stats['failing_count']]
        ]
        
        for i, (metric, value) in enumerate(stats_data, 6):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
            
        # تنسيق الجدول
        for row in range(6, 16):
            for col in ['A', 'B']:
                cell = ws[f'{col}{row}']
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # ضبط عرض الأعمدة
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        
    def _create_detailed_data_sheet(self, wb: openpyxl.Workbook, df: pd.DataFrame):
        """إنشاء ورقة البيانات التفصيلية"""
        ws = wb.create_sheet("البيانات التفصيلية")
        
        # إضافة عمود التصنيف والحالة باستخدام النسبة المئوية إن وجدت
        detailed_df = df.copy()
        
        if 'النسبة المئوية' in detailed_df.columns:
            # استخدام النسبة المئوية للتصنيف والحالة
            detailed_df['التصنيف'] = detailed_df['النسبة المئوية'].apply(self._classify_grade)
            detailed_df['الحالة'] = detailed_df['النسبة المئوية'].apply(lambda x: 'ناجح' if x >= 50 else 'راسب')
            detailed_df = detailed_df.sort_values('النسبة المئوية', ascending=False).reset_index(drop=True)
        else:
            # استخدام الدرجة المطلقة (الطريقة التقليدية)
            detailed_df['التصنيف'] = detailed_df['الدرجة'].apply(self._classify_grade)
            detailed_df['الحالة'] = detailed_df['الدرجة'].apply(lambda x: 'ناجح' if x >= self.passing_grade else 'راسب')
            detailed_df = detailed_df.sort_values('الدرجة', ascending=False).reset_index(drop=True)
            
        detailed_df.index = detailed_df.index + 1
        
        # العنوان
        if 'النسبة المئوية' in df.columns:
            ws.merge_cells('A1:G1')
        else:
            ws.merge_cells('A1:F1')
        ws['A1'] = "البيانات التفصيلية للطلاب"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        
        # إضافة البيانات
        headers = ['الترتيب', 'اسم الطالب', 'الدرجة', 'النسبة المئوية', 'التصنيف', 'الحالة']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # إضافة البيانات
        for row_idx, (_, row_data) in enumerate(detailed_df.iterrows(), 4):
            ws.cell(row=row_idx, column=1, value=row_data.name)
            ws.cell(row=row_idx, column=2, value=row_data['اسم الطالب'])
            ws.cell(row=row_idx, column=3, value=row_data['الدرجة'])
            
            # النسبة المئوية
            if 'النسبة المئوية' in detailed_df.columns:
                ws.cell(row=row_idx, column=4, value=f"{row_data['النسبة المئوية']:.1f}%")
            else:
                ws.cell(row=row_idx, column=4, value="غير محسوبة")
            
            ws.cell(row=row_idx, column=5, value=row_data['التصنيف'])
            
            # تلوين حالة الطالب
            status_cell = ws.cell(row=row_idx, column=6, value=row_data['الحالة'])
            if row_data['الحالة'] == 'ناجح':
                status_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            else:
                status_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # تنسيق الجدول
        for row in range(3, len(detailed_df) + 4):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center')
        
        # ضبط عرض الأعمدة
        column_widths = [10, 25, 10, 15, 15, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_grade_ranges_sheet(self, wb: openpyxl.Workbook, grade_ranges: pd.DataFrame):
        """إنشاء ورقة نطاقات الدرجات"""
        ws = wb.create_sheet("نطاقات الدرجات")
        
        # العنوان
        ws.merge_cells('A1:C1')
        ws['A1'] = "توزيع الطلاب حسب نطاقات الدرجات"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        
        # إضافة البيانات
        for r in dataframe_to_rows(grade_ranges, index=False, header=True):
            ws.append(r)
        
        # تنسيق الرؤوس
        for col in range(1, 4):
            cell = ws.cell(row=3, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # تنسيق البيانات
        for row in range(3, len(grade_ranges) + 4):
            for col in range(1, 4):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center')
        
        # ضبط عرض الأعمدة
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _create_top_students_sheet(self, wb: openpyxl.Workbook, df: pd.DataFrame):
        """إنشاء ورقة الطلاب المتفوقين"""
        ws = wb.create_sheet("الطلاب المتفوقين")
        
        top_students = df.nlargest(10, 'الدرجة').reset_index(drop=True)
        
        # العنوان
        ws.merge_cells('A1:C1')
        ws['A1'] = "أفضل 10 طلاب"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        
        # الرؤوس
        if 'النسبة المئوية' in df.columns:
            headers = ['المرتبة', 'اسم الطالب', 'الدرجة', 'النسبة المئوية']
        else:
            headers = ['المرتبة', 'اسم الطالب', 'الدرجة']
            
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # البيانات
        for row_idx, (_, student) in enumerate(top_students.iterrows(), 4):
            ws.cell(row=row_idx, column=1, value=row_idx - 3)
            ws.cell(row=row_idx, column=2, value=student['اسم الطالب'])
            ws.cell(row=row_idx, column=3, value=student['الدرجة'])
            
            # إضافة النسبة المئوية إن وجدت
            if 'النسبة المئوية' in df.columns:
                ws.cell(row=row_idx, column=4, value=f"{student['النسبة المئوية']:.1f}%")
        
        # تنسيق الجدول
        col_count = 4 if 'النسبة المئوية' in df.columns else 3
        for row in range(3, len(top_students) + 4):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center')
        
        # ضبط عرض الأعمدة
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 10
        if 'النسبة المئوية' in df.columns:
            ws.column_dimensions['D'].width = 15
    
    def _create_failing_students_sheet(self, wb: openpyxl.Workbook, df: pd.DataFrame):
        """إنشاء ورقة الطلاب المتعثرين"""
        ws = wb.create_sheet("الطلاب المتعثرين")
        
        # استخدام النسبة المئوية للتقييم إن وجدت
        if 'النسبة المئوية' in df.columns:
            failing_students = df[df['النسبة المئوية'] < 50].copy()
            # حساب الفجوة: 50% من الدرجة الكلية ناقص درجة الطالب
            if 'الدرجة الكلية' in df.columns:
                failing_students['الفجوة'] = (failing_students['الدرجة الكلية'] * 0.5) - failing_students['الدرجة']
            else:
                failing_students['الفجوة'] = 25 - failing_students['الدرجة']  # افتراض أن الدرجة من 50
            failing_students = failing_students.sort_values('الدرجة', ascending=True).reset_index(drop=True)
        else:
            failing_students = df[df['الدرجة'] < self.passing_grade].copy()
            failing_students['الفجوة'] = self.passing_grade - failing_students['الدرجة']
            failing_students = failing_students.sort_values('الدرجة', ascending=True).reset_index(drop=True)
        
        # العنوان
        ws.merge_cells('A1:D1')
        ws['A1'] = "الطلاب المتعثرين (نسبة أقل من 50%)"
            
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].fill = PatternFill(start_color='D32F2F', end_color='D32F2F', fill_type='solid')
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        
        # الرؤوس
        headers = ['اسم الطالب', 'الدرجة', 'الفجوة', 'ملاحظات']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # البيانات
        for row_idx, (_, student) in enumerate(failing_students.iterrows(), 4):
            ws.cell(row=row_idx, column=1, value=student['اسم الطالب'])
            ws.cell(row=row_idx, column=2, value=student['الدرجة'])
            ws.cell(row=row_idx, column=3, value=f"{student['الفجوة']:.1f}")
            
            # ملاحظات حسب النسبة المئوية أو الدرجة
            if 'النسبة المئوية' in df.columns:
                grade_for_note = student['النسبة المئوية']
            else:
                grade_for_note = student['الدرجة']
                
            if grade_for_note < 30:
                note = "يحتاج دعم عاجل"
            elif grade_for_note < 40:
                note = "يحتاج تدخل سريع"
            else:
                note = "قريب من النجاح"
            
            ws.cell(row=row_idx, column=4, value=note)
        
        # تنسيق الجدول
        for row in range(3, len(failing_students) + 4):
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal='center')
        
        # ضبط عرض الأعمدة
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 20
    
    def _create_statistics_sheet(self, wb: openpyxl.Workbook, df: pd.DataFrame, stats: Dict):
        """إنشاء ورقة الإحصائيات المتقدمة"""
        ws = wb.create_sheet("الإحصائيات المتقدمة")
        
        # العنوان
        ws.merge_cells('A1:B1')
        ws['A1'] = "الإحصائيات المتقدمة"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        
        grades = df['الدرجة']
        
        # الإحصائيات المتقدمة
        advanced_stats = [
            ['الإحصائية', 'القيمة'],
            ['المتوسط الحسابي', f"{grades.mean():.3f}"],
            ['المتوسط الهندسي', f"{grades.apply(lambda x: np.log(x) if x > 0 else 0).mean():.3f}"],
            ['الوسيط', f"{grades.median():.3f}"],
            ['المنوال', f"{grades.mode().iloc[0] if not grades.mode().empty else 'غير محدد'}"],
            ['الانحراف المعياري', f"{grades.std():.3f}"],
            ['التباين', f"{grades.var():.3f}"],
            ['معامل الاختلاف', f"{(grades.std()/grades.mean())*100:.2f}%"],
            ['الالتواء (Skewness)', f"{grades.skew():.3f}"],
            ['التفلطح (Kurtosis)', f"{grades.kurtosis():.3f}"],
            ['الربع الأول (Q1)', f"{grades.quantile(0.25):.2f}"],
            ['الربع الثالث (Q3)', f"{grades.quantile(0.75):.2f}"],
            ['المدى الربعي (IQR)', f"{grades.quantile(0.75) - grades.quantile(0.25):.2f}"],
            ['المدى', f"{grades.max() - grades.min():.2f}"],
            ['الحد الأدنى للقيم الشاذة', f"{grades.quantile(0.25) - 1.5 * (grades.quantile(0.75) - grades.quantile(0.25)):.2f}"],
            ['الحد الأعلى للقيم الشاذة', f"{grades.quantile(0.75) + 1.5 * (grades.quantile(0.75) - grades.quantile(0.25)):.2f}"]
        ]
        
        for row_idx, (stat, value) in enumerate(advanced_stats, 3):
            ws.cell(row=row_idx, column=1, value=stat)
            ws.cell(row=row_idx, column=2, value=value)
            
            if row_idx == 3:  # الرأس
                for col in [1, 2]:
                    cell = ws.cell(row=row_idx, column=col)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center')
        
        # تنسيق الجدول
        for row in range(3, len(advanced_stats) + 3):
            for col in [1, 2]:
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # ضبط عرض الأعمدة
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
    
    def _classify_grade(self, grade: float) -> str:
        """تصنيف الدرجة حسب النطاق"""
        if grade >= 90:
            return 'ممتاز'
        elif grade >= 80:
            return 'جيد جداً'
        elif grade >= 70:
            return 'جيد'
        elif grade >= 60:
            return 'مقبول'
        elif grade >= 50:
            return 'ضعيف'
        else:
            return 'راسب'
